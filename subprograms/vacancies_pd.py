from openpyxl import Workbook
from openpyxl.styles import Border, Side, Font
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
from math import floor
from jinja2 import Environment, FileSystemLoader
import pdfkit

file_name = input("Введите название файла: ")
vacancy_name = input("Введите название профессии: ")

df = pd.read_csv(file_name, encoding="utf-8")
q = df['salary'].quantile(0.99)
df = df[df['salary'] < q]

def split_years(dt):
    return [dt[dt['published_at'].str[:4] == y] for y in dt['published_at'].str[:4].unique()]

res = split_years(df)

res_selected_vacancy = [i[i["name"].str.contains(vacancy_name)] for i in res]

vacancies_average_salary_by_year = {}
vacancies_count_by_year = {}
for i in res:
    vacancies_average_salary_by_year[int(i['published_at'].iat[0][:4])] = floor(i['salary'].mean())
    vacancies_count_by_year[int(i['published_at'].iat[0][:4])] = len(i.index)

vacancies_average_salary_by_year_selected_name = {}
vacancies_count_by_year_selected_name = {}
for i in res_selected_vacancy:
    vacancies_average_salary_by_year_selected_name[int(i['published_at'].iat[0][:4])] = floor(i['salary'].mean())
    vacancies_count_by_year_selected_name[int(i['published_at'].iat[0][:4])] = len(i.index)

print(vacancies_average_salary_by_year)
print(vacancies_count_by_year)
print(vacancies_average_salary_by_year_selected_name)
print(vacancies_count_by_year_selected_name)


border = Border(left=Side(border_style='thin', color='000000'),
                right=Side(border_style='thin', color='000000'),
                top=Side(border_style='thin', color='000000'),
                bottom=Side(border_style='thin', color='000000'))
font = Font(bold=True)

def generate_excel():
    wb = Workbook()
    ws = wb.active
    ws.title = "Статистика по годам"
    fill_first_table(ws)
    set_correct_cell_width(ws)
    return ws


def fill_first_table(ws):
    """Заполняет первый лист таблицы.
    Args:
        ws (openpyxl.workbook.workbook.Workbook @property): лист таблицы
    """
    set_first_page_headers(ws)
    years = list(vacancies_average_salary_by_year.keys())
    for row in range(0, len(years)):
        set_cell_border(ws.cell(column=1, row=row + 2, value=years[row]))
        set_cell_border(
            ws.cell(column=2, row=row + 2, value=vacancies_average_salary_by_year[years[row]]))
        set_cell_border(
            ws.cell(column=3, row=row + 2, value=vacancies_average_salary_by_year_selected_name[years[row]]))
        set_cell_border(ws.cell(column=4, row=row + 2, value=vacancies_count_by_year[years[row]]))
        set_cell_border(
            ws.cell(column=5, row=row + 2, value=vacancies_count_by_year_selected_name[years[row]]))

def set_correct_cell_width( ws):
    """Форматирует ширину клеток таблицы.
    Args:
        ws (openpyxl.workbook.workbook.Workbook @property): лист таблицы
    """
    dims = {}
    for row in ws.rows:
        for cell in row:
            if cell.value:
                dims[cell.column_letter] = max((dims.get(cell.column_letter, 0), len(str(cell.value))))
    for col, value in dims.items():
        ws.column_dimensions[col].width = value + 2

def set_cell_border(cell):
    """Устанавливает границу клетки таблицы.

    Args:
        cell ((openpyxl.workbook.workbook.Workbook @property).cell): Клетка таблицы
    """
    cell.border = border

def set_first_page_headers( ws):
    """Устанавливает заголовки первого листа таблицы.

    Args:
        ws (openpyxl.workbook.workbook.Workbook @property): Лист таблицы
    """
    ws["A1"] = "Год"
    ws["B1"] = "Средняя зарплата"
    ws["C1"] = f"Средняя зарплата - {vacancy_name}"
    ws["D1"] = "Количество вакансий"
    ws["E1"] = f"Количество вакансий - {vacancy_name}"
    for row in ws.iter_rows(max_row=1, max_col=5):
        for cell in row:
            set_cell_border(cell)
            cell.font = font

def generate_image():
        """Генерирует изображение графиков"""
        fig, axes = plt.subplots(nrows=1, ncols=2)
        generate_first_plot(axes[0])
        generate_second_plot(axes[1])
        plt.tight_layout()
        plt.savefig(r"D:\PythonProject\report\pd_graph.png", dpi=200)

def generate_first_plot(firstplt):
    """Генерирует первый график

    Args: firstplt (axis): Первый график
    """
    labels = list(vacancies_average_salary_by_year.keys())
    average_salary_year = list(vacancies_average_salary_by_year.values())
    average_salary_year_selected = list(vacancies_average_salary_by_year_selected_name.values())
    x = np.array(labels)
    width = 0.4
    firstplt.set_title("Уровень зарплат по годам")
    firstplt.set_xticks(x)
    firstplt.tick_params(axis='x', rotation=90, labelsize=8)
    firstplt.set_yticks(np.arange(0, max(average_salary_year_selected), step=20000))
    firstplt.tick_params(axis='y', labelsize=8)
    firstplt.grid(axis="y")
    firstplt.bar(x - width / 2, average_salary_year, width, label="средняя з/п")
    firstplt.bar(x + width / 2, average_salary_year_selected, width, label=f"з/п {vacancy_name}")
    firstplt.legend(fontsize=8)

def generate_second_plot( secondplt):
    """Генерирует второй график

    Args: secondplt (axis) : второй график
    """
    labels = list(vacancies_count_by_year.keys())
    count_by_year = list(vacancies_count_by_year.values())
    count_by_year_selected = list(vacancies_count_by_year_selected_name.values())
    x = np.array(labels)
    y = np.array(count_by_year)
    width = 0.4
    secondplt.set_title("Количество вакансий по годам")
    secondplt.set_xticks(x)
    secondplt.tick_params(axis='x', rotation=90, labelsize=8)
    secondplt.set_yticks(np.arange(0, max(count_by_year), step=20000))
    secondplt.tick_params(axis='y', labelsize=8)
    secondplt.grid(axis='y')
    secondplt.bar(x - width / 2, count_by_year, width, label="Количество вакансий")
    secondplt.bar(x + width / 2, count_by_year_selected, width, label=f"Количество вакансий\n{vacancy_name}")
    secondplt.legend(fontsize=8)

generate_image()

def generate_pdf():
    """Генерирует pdf файл"""
    env = Environment(loader=FileSystemLoader('.'))
    template = env.get_template("pdf_template_pd.html")
    sheet = generate_excel()
    pdf_template = template.render({'vacancy_name': vacancy_name}, sheet=sheet)
    config = pdfkit.configuration(wkhtmltopdf=r'D:\wkhtmltox\bin\wkhtmltopdf.exe')
    pdfkit.from_string(pdf_template, 'report_pd.pdf', configuration=config, options={"enable-local-file-access": None})

generate_pdf()