import csv
import functools
import re
from openpyxl import Workbook
from openpyxl.reader.excel import load_workbook
from openpyxl.styles import Border, Side, Font
from datetime import datetime
from math import floor
from statistics import mean
import matplotlib.pyplot as plt
import numpy as np
from openpyxl.styles.numbers import FORMAT_PERCENTAGE_00
from jinja2 import Environment, FileSystemLoader
import pdfkit
from unittest import TestCase
import cProfile
from multiprocessing import Pool
import os
import glob
from collections import Counter
from functools import reduce
from itertools import groupby


class Salary:
    """Класс для представления зарплаты.

    Attributes:
        salary_from (int or float or str): Нижняя граница вилки оклада
        salary_to (int or float or str): Верхняя граница вилки оклада
        salary_currency (str): Валюта оклада
        rub_salary (float): Средняя зарплата, переведенная в рубли с помощью словаря currency_to_rub
    """
    currency_to_rub = {
        "AZN": 35.68,
        "BYR": 23.91,
        "EUR": 59.90,
        "GEL": 21.74,
        "KGS": 0.76,
        "KZT": 0.13,
        "RUR": 1,
        "UAH": 1.64,
        "USD": 60.66,
        "UZS": 0.0055,
    }

    def __init__(self, salary_from, salary_to, salary_currency):
        """Инициализирует объект Salary, выполняет конвертацию для целочисленных полей.

        Args:
            salary_from (int or float or str): Нижняя граница вилки оклада
            salary_to (int or float or str): Верхняя граница вилки оклада
            salary_currency (str): Валюта оклада

        >>> type(Salary(10.0, 20.4, 'RUR')).__name__
        'Salary'
        >>> Salary(10.0, 20.4, 'RUR').salary_from
        10
        >>> Salary(10.0, 20.4, 'RUR').salary_to
        20
        >>> Salary(10.0, 20.4, 'RUR').salary_currency
        'RUR'
        """
        self.salary_from = int(float(salary_from))
        self.salary_to = int(float(salary_to))
        self.salary_currency = salary_currency
        self.rub_salary = self.get_salary_in_rub()

    def get_salary_in_rub(self):
        """Вычисляет среднюю зарплату из вилки и переводит в рубли, при помощи словаря

        Returns:
            float: Средняя зарплата в рублях
        """
        return ((self.salary_from + self.salary_to) / 2) * self.currency_to_rub[self.salary_currency]


class SalaryTests(TestCase):
    def test_salary_type(self):
        self.assertEqual(type(Salary(10.0, 20.4, 'RUR')).__name__, 'Salary')

    def test_salary_from(self):
        self.assertEqual(Salary(10.0, 20.4, 'RUR').salary_from, 10)

    def test_salary_to(self):
        self.assertEqual(Salary(10.0, 20.4, 'RUR').salary_to, 20)

    def test_salary_currency(self):
        self.assertEqual(Salary(10.0, 20.4, 'RUR').salary_currency, 'RUR')

    def test_int_get_salary(self):
        self.assertEqual(Salary(10, 20, 'RUR').get_salary_in_rub(), 15.0)

    def test_float_salary_from_in_get_salary(self):
        self.assertEqual(Salary(10.0, 20, 'RUR').get_salary_in_rub(), 15.0)

    def test_float_salary_to_in_get_salary(self):
        self.assertEqual(Salary(10, 20.0, 'RUR').get_salary_in_rub(), 15.0)

    def test_currency_in_get_salary(self):
        self.assertEqual(Salary(10, 30.0, 'EUR').get_salary_in_rub(), 1198.0)


class Vacancy:
    """Класс для хранения данных о вакансии.

    Attributes:
        name (str): Название вакансии
        description (str): Описание вакансии
        key_skills (str): Навыки
        experience_id (str): Опыт работы
        premium (str): Премиум вакансия
        employer_name (str): Компания
        salary (Salary): Зарплата в рублях
        area_name (str): Город
        published_at (str): Дата публикации вакансии
    """

    def __init__(self, vacancy_dictionary):
        """Иницилиазирует объект Vacancy из словаря vacancy_dictionary.

        Args:
            vacancy_dictionary (dict): Словарь с исходными данными о вакансии
        """
        self.name = vacancy_dictionary.get("name")
        self.description = vacancy_dictionary.get("description")
        self.key_skills = vacancy_dictionary.get("key_skills")
        self.experience_id = vacancy_dictionary.get("experience_id")
        self.premium = vacancy_dictionary.get("premium")
        self.employer_name = vacancy_dictionary.get("employer_name")
        try:
            self.salary = Salary(vacancy_dictionary.get("salary_from"),
                                 vacancy_dictionary.get("salary_to"),
                                 vacancy_dictionary.get("salary_currency"))
        except:
            self.salary = None
        self.area_name = vacancy_dictionary.get("area_name")
        self.published_at = vacancy_dictionary.get("published_at")

    def parse_datetime(self):
        """Парсер даты.

        Returns:
            datetime: Дата полученная с помощью strptime
        """
        return datetime.strptime(self.published_at, "%Y-%m-%dT%H:%M:%S%z")


class VacancyTests(TestCase):
    def test_parse_datetime_strptime_type(self):
        self.assertEqual(type(Vacancy({"published_at": "2007-12-03T17:47:55+0300"})
                              .parse_datetime()).__name__, "datetime")

    def test_parse_datetime_strptime_year(self):
        self.assertEqual(Vacancy({"published_at": "2007-12-03T17:47:55+0300"})
                         .parse_datetime().year, 2007)

    def test_parse_datetime_strptime_day(self):
        self.assertEqual(Vacancy({"published_at": "2007-12-03T17:47:55+0300"})
                         .parse_datetime().day, 3)

    def test_parse_datetime_strptime_month(self):
        self.assertEqual(Vacancy({"published_at": "2007-12-03T17:47:55+0300"})
                         .parse_datetime().month, 12)

    def test_Salary_type(self):
        self.assertEqual(type
                         (Vacancy({"salary_from": 0, "salary_to": 0, "salary_currency": "RUR"}).salary).__name__,
                         "Salary")

    def test_if_Salary_is_None_when_empty_salary_keys(self):
        self.assertEqual(type(Vacancy({}).salary).__name__, "NoneType")

    def test_if_Salary_is_None_when_empty_salary_from(self):
        self.assertEqual(type(Vacancy({"salary_to": 1, "salary_currency": 'RUR'}).salary).__name__, "NoneType")

    def test_if_Salary_is_None_when_empty_salary_to(self):
        self.assertEqual(type(Vacancy({"salary_from": 1, "salary_currency": 'RUR'}).salary).__name__, "NoneType")

    def test_if_Salary_is_None_when_empty_salary_currency(self):
        self.assertEqual(type(Vacancy({"salary_to": 1, "salary_from": 1}).salary).__name__, "NoneType")


class DataSet:
    """ Класс для обработки входных данных в формате CSV.

    Attributes:
        file_name (str): Название файла в формате "*****.csv"
        vacancies_objects (list): Список вакансий в виде объектов обработанных csv парсером.
    """

    def __init__(self, file_name):
        """Инициализирует Dataset, выполняет парсинг CSV файла

        Args:
            file_name (str): Название файла в формате "*****.csv"
        """
        self.file_name = file_name
        self.vacancies_objects = self.csv_OOP_parser(file_name)

    @staticmethod
    def remove_html(string):
        """Убирает html из строки.

        Args:
             string(str): Строка

        Returns:
            string: Строка, очищенная от HTML символов

        >>> DataSet.remove_html("<p>123</p>")
        '123'
        >>> DataSet.remove_html("<div>divdiv</div>")
        'divdiv'
        >>> DataSet.remove_html("<p>")
        ''
        >>> DataSet.remove_html(" <p> ")
        '  '
        >>> DataSet.remove_html("")
        ''
        >>> DataSet.remove_html("abc")
        'abc'
        """
        regex = re.compile(r'<[^>]+>')
        return regex.sub('', string)

    @staticmethod
    def repair_string(s):
        """Убирает из строки лишние пробелы и HTML символы.

        Args:
            s (str): Строка

        Returns:
            string: Строка, без лишних пробелов и HTML символов
        >>> DataSet.repair_string("<p> Raw </p>")
        'Raw'
        >>> DataSet.repair_string(" ")
        ''
        >>> DataSet.repair_string("<p>  </p>")
        ''
        """
        s = DataSet.remove_html(s)
        s = s.strip()
        s = " ".join(s.split())
        return s

    @staticmethod
    def fill_vacancy_dictionary(field_names, row):
        """Заполняет словарь вакансии, ключами которого являются названия входных полей.

        Args:
            field_names (list(str)) : Список полей вакансии, полученный из заголовка таблицы
            row (list(str)): Ряд таблицы, содержащий данные о вакансии

        Returns:
            dict: Словарь с ключами-заголовками field_names и соответствующими им полями вакансии

        >>> DataSet.fill_vacancy_dictionary(["name"], ["Олег"])
        {'name': 'Олег'}
        >>> DataSet.fill_vacancy_dictionary([],[])
        {}
        >>> DataSet.fill_vacancy_dictionary(['name', 'salary_from'], ['Олег'])
        0
        >>> DataSet.fill_vacancy_dictionary([], ["Олег"])
        0
        >>> DataSet.fill_vacancy_dictionary(['name'], [''])
        0
        """
        vacancy_dictionary = {}
        if len(row) != len(field_names):
            return 0
        for i in range(len(field_names)):
            if len(row[i]) == 0:
                return 0
            DataSet.repair_string(row[i])
            vacancy_dictionary[field_names[i]] = row[i]
        return vacancy_dictionary

    def slice_parse_year(self, date_string):
        return date_string[0:4]

    def csv_OOP_parser(self, file_name):
        """Парсер csv файла.

        Args:
            file_name (str): Название файла в формате "*****.csv"

        Returns:
            list(Vacancy): Список вакансий в виде объектов
        """
        f = open(file_name, mode="r", encoding="utf-8-sig")
        reader = csv.reader(f)
        try:
            field_names = next(reader)
        except:
            return 0

        vacancies_objects = []
        for row in reader:
            if len(row) < len(field_names):
                continue
            vacancy_dict = self.fill_vacancy_dictionary(field_names, row)
            if vacancy_dict == 0:
                continue
            vacancy = Vacancy(vacancy_dict)
            vacancies_objects.append(vacancy)
        return vacancies_objects

    def process_vacancies(self, vacancy_name):
        """Обрабатывает вакансии и возвращает данные о них по категориям.

        Args:
            vacancy_name (str): Название вакансии

        Returns:
            vacancies_average_salary_by_year (dict): Динамика уровня зарплат по годам
            vacancies_count_by_year (dict): Динамика количества вакансий по годам
            vacancies_average_salary_by_year_selected_name (dict): Динамика уровня зарплат по годам
                                                                                                для выбранной профессии
            vacancies_count_by_year_selected_name (dict): Динамика количества вакансий по годам для выбранной профессии
            vacancies_top_average_salary_by_city (dict): Уровень зарплат по городам (в порядке убывания) - только
                                                                                                    первые 10 значений
            vacancies_fraction_by_city (dict): Доля вакансий по городам
            vacancy_name (str): Название вакансии
        """
        vacancies_average_salary_by_year = {}
        vacancies_count_by_year = {}
        vacancies_average_salary_by_year_selected_name = {}
        vacancies_count_by_year_selected_name = {}
        vacancies_average_salary_by_city = {}
        vacancies_count_by_city = {}
        vacancies_objects = self.vacancies_objects
        for vacancy in vacancies_objects:
            year = self.slice_parse_year(vacancy.published_at)
            if vacancy.area_name in vacancies_count_by_city:
                vacancies_count_by_city[vacancy.area_name] += 1
            else:
                vacancies_count_by_city[vacancy.area_name] = 1

            if year in vacancies_count_by_year:
                vacancies_count_by_year[year] += 1
            else:
                vacancies_count_by_year[year] = 1
                vacancies_count_by_year_selected_name[year] = 0

            if year in vacancies_average_salary_by_year:
                vacancies_average_salary_by_year[year].append(vacancy.salary.rub_salary)
            else:
                vacancies_average_salary_by_year[year] = [vacancy.salary.rub_salary]
                vacancies_average_salary_by_year_selected_name[year] = 0

            if vacancy.area_name in vacancies_average_salary_by_city:
                vacancies_average_salary_by_city[vacancy.area_name].append(vacancy.salary.rub_salary)
            else:
                vacancies_average_salary_by_city[vacancy.area_name] = [vacancy.salary.rub_salary]

            if vacancy_name in vacancy.name:
                if year in vacancies_count_by_year_selected_name:
                    vacancies_count_by_year_selected_name[year] += 1

                if vacancies_average_salary_by_year_selected_name[year] == 0:
                    vacancies_average_salary_by_year_selected_name[year] = [vacancy.salary.rub_salary]
                else:
                    vacancies_average_salary_by_year_selected_name[year].append(vacancy.salary.rub_salary)
        vacancies_salary_sum_by_year = {}
        
        for key in vacancies_average_salary_by_year:
            vacancies_average_salary_by_year[key] = int(mean(vacancies_average_salary_by_year[key]))
        for key in vacancies_average_salary_by_year_selected_name:
            if vacancies_average_salary_by_year_selected_name[key] != 0:
                vacancies_average_salary_by_year_selected_name[key] = int(
                    mean(vacancies_average_salary_by_year_selected_name[key]))
        for key in vacancies_average_salary_by_city:
            vacancies_average_salary_by_city[key] = int(mean(vacancies_average_salary_by_city[key]))

        city_count = sum(vacancies_count_by_city.values())
        one_percent = floor(city_count / 100)

        vacancies_fraction_by_city = {k: round(v / city_count, 4)
                                      for k, v in sorted(vacancies_count_by_city.items(),
                                                         reverse=True,
                                                         key=lambda item: item[1])
                                      if v >= one_percent}
        vacancies_top_average_salary_by_city = {k: v
                                                for k, v in sorted(vacancies_average_salary_by_city.items(),
                                                                   reverse=True,
                                                                   key=lambda item: item[1])
                                                if vacancies_count_by_city[k] >= one_percent}

        return vacancies_average_salary_by_year, \
               vacancies_count_by_year, \
               vacancies_average_salary_by_year_selected_name, \
               vacancies_count_by_year_selected_name, \
               vacancies_top_average_salary_by_city, \
               vacancies_fraction_by_city, \
               vacancy_name, \
               vacancies_count_by_city, \
                vacancies_average_salary_by_city 
               


# file_name = input("Введите название файла: ")
# vacancy_name = input("Введите название профессии: ")
#
# dataset = DataSet(file_name)
# processed_data = dataset.process_vacancies(vacancy_name)
#
# print(f"Динамика уровня зарплат по годам: {processed_data[0]}")
# print(f"Динамика количества вакансий по годам: {processed_data[1]}")
# print(f"Динамика уровня зарплат по годам для выбранной профессии: {processed_data[2]}")
# print(f"Динамика количества вакансий по годам для выбранной профессии: {processed_data[3]}")
# print(f"Уровень зарплат по городам (в порядке убывания): {dict(list(processed_data[4].items())[:10])}")
# print(f"Доля вакансий по городам (в порядке убывания): {dict(list(processed_data[5].items())[:10])}")


class Report:
    """Класс для формирования отчета в виде таблицы, изображения или PDF файла.

    Attributes:
            vacancies_average_salary_by_year (dict): Динамика уровня зарплат по годам
            vacancies_count_by_year (dict): Динамика количества вакансий по годам
            vacancies_average_salary_by_year_selected_name (dict): Динамика уровня зарплат по годам
                                                                                                для выбранной профессии
            vacancies_count_by_year_selected_name (dict): Динамика количества вакансий по годам для выбранной профессии
            vacancies_top_average_salary_by_city (dict): Уровень зарплат по городам (в порядке убывания) - только
                                                                                                    первые 10 значений
            vacancies_fraction_by_city (dict): Доля вакансий по городам
            vacancy_name (str): Название вакансии
    """
    border = Border(left=Side(border_style='thin', color='000000'),
                    right=Side(border_style='thin', color='000000'),
                    top=Side(border_style='thin', color='000000'),
                    bottom=Side(border_style='thin', color='000000'))
    font = Font(bold=True)

    def __init__(self, processed_data):
        """Инициализирует данные для формирования отчета.

        Args:
            processed_data (list): Список информации о вакансиях полученных из DataSet
        """
        self.vacancies_average_salary_by_year = processed_data[0]
        self.vacancies_count_by_year = processed_data[1]
        self.vacancies_average_salary_by_year_selected_name = processed_data[2]
        self.vacancies_count_by_year_selected_name = processed_data[3]
        self.vacancies_top_average_salary_by_city = dict(list(processed_data[4].items())[:10])
        self.vacancies_fraction_by_city = dict(list(processed_data[5].items())[:10])
        self.vacancy_name = processed_data[6]

    def generate_excel(self):
        """Генерирует excel файл в папке с файлом"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Статистика по годам"
        ws2 = wb.create_sheet("Статистика по городам")
        self.fill_first_table(ws)
        self.set_correct_cell_width(ws)
        self.fill_second_table(ws2)
        self.set_correct_cell_width(ws2)
        ws2.column_dimensions['C'].width = 2
        wb.save("report.xlsx")

    def fill_first_table(self, ws):
        """Заполняет первый лист таблицы.
        Args:
            ws (openpyxl.workbook.workbook.Workbook @property): лист таблицы
        """
        self.set_first_page_headers(ws)
        years = list(self.vacancies_average_salary_by_year.keys())
        for row in range(0, len(years)):
            self.set_cell_border(ws.cell(column=1, row=row + 2, value=years[row]))
            self.set_cell_border(
                ws.cell(column=2, row=row + 2, value=self.vacancies_average_salary_by_year[years[row]]))
            self.set_cell_border(
                ws.cell(column=3, row=row + 2, value=self.vacancies_average_salary_by_year_selected_name[years[row]]))
            self.set_cell_border(ws.cell(column=4, row=row + 2, value=self.vacancies_count_by_year[years[row]]))
            self.set_cell_border(
                ws.cell(column=5, row=row + 2, value=self.vacancies_count_by_year_selected_name[years[row]]))

    def fill_second_table(self, ws2):
        """Заполняет второй лист таблицы.
        Args:
            ws2 (openpyxl.workbook.workbook.Workbook @property): лист таблицы
        """
        self.set_second_page_headers(ws2)
        city_list_top_salary = list(self.vacancies_top_average_salary_by_city)
        for row in range(0, len(self.vacancies_top_average_salary_by_city)):
            self.set_cell_border(ws2.cell(column=1, row=row + 2, value=city_list_top_salary[row]))
            self.set_cell_border(ws2.cell(column=2, row=row + 2,
                                          value=self.vacancies_top_average_salary_by_city[city_list_top_salary[row]]))
        city_list_fraction = list(self.vacancies_fraction_by_city)
        for row in range(0, len(self.vacancies_top_average_salary_by_city)):
            self.set_cell_border(ws2.cell(column=4, row=row + 2, value=city_list_fraction[row]))
            cell = ws2.cell(column=5, row=row + 2, value=self.vacancies_fraction_by_city[city_list_fraction[row]])
            cell.number_format = FORMAT_PERCENTAGE_00
            self.set_cell_border(cell)

    def set_correct_cell_width(self, ws):
        """Форматирует ширину клеток таблицы.
        Args:
            ws (openpyxl.workbook.workbook.Workbook @property): лист таблицы
        """
        dims = {}
        for row in ws.rows:
            for cell in row:
                if cell.value:
                    dims[cell.column_letter] = max((dims.get(cell.column_letter, 0), len(str(cell.value))))
        for col, value in dims.items():
            ws.column_dimensions[col].width = value + 2

    def set_cell_border(self, cell):
        """Устанавливает границу клетки таблицы.

        Args:
            cell ((openpyxl.workbook.workbook.Workbook @property).cell): Клетка таблицы
        """
        cell.border = self.border

    def set_first_page_headers(self, ws):
        """Устанавливает заголовки первого листа таблицы.

        Args:
            ws (openpyxl.workbook.workbook.Workbook @property): Лист таблицы
        """
        ws["A1"] = "Год"
        ws["B1"] = "Средняя зарплата"
        ws["C1"] = f"Средняя зарплата - {self.vacancy_name}"
        ws["D1"] = "Количество вакансий"
        ws["E1"] = f"Количество вакансий - {self.vacancy_name}"
        for row in ws.iter_rows(max_row=1, max_col=5):
            for cell in row:
                self.set_cell_border(cell)
                cell.font = self.font

    def set_second_page_headers(self, ws2):
        """Устанавливает заголовки второго листа таблицы

        Args:
            ws2 (openpyxl.workbook.workbook.Workbook @property): Лист таблицы
        """
        ws2["A1"] = "Город"
        ws2["A1"].font = self.font
        ws2["A1"].border = self.border

        ws2["B1"] = "Уровень зарплат"
        ws2["B1"].font = self.font
        ws2["B1"].border = self.border

        ws2["D1"] = "Город"
        ws2["D1"].font = self.font
        ws2["D1"].border = self.border

        ws2["E1"] = "Доля вакансий"
        ws2["E1"].font = self.font
        ws2["E1"].border = self.border

    def generate_image(self):
        """Генерирует изображение графиков"""
        fig, axes = plt.subplots(nrows=2, ncols=2)
        self.generate_first_plot(axes[0, 0])
        self.generate_second_plot(axes[0, 1])
        self.generate_third_plot(axes[1, 0])
        self.generate_fourth_plot(axes[1, 1])
        plt.savefig(r"D:\PythonProject\report\graph.png", dpi=200)

    def generate_first_plot(self, firstplt):
        """Генерирует первый график

        Args: firstplt (axis): Первый график
        """
        labels = list(self.vacancies_average_salary_by_year.keys())
        average_salary_year = list(self.vacancies_average_salary_by_year.values())
        average_salary_year_selected = list(self.vacancies_average_salary_by_year_selected_name.values())
        x = np.array(labels)
        width = 0.4
        firstplt.set_title("Уровень зарплат по годам")
        firstplt.set_xticks(x)
        firstplt.tick_params(axis='x', rotation=90, labelsize=8)
        firstplt.set_yticks(np.arange(0, max(average_salary_year_selected), step=10000))
        firstplt.tick_params(axis='y', labelsize=8)
        firstplt.grid(axis="y")
        firstplt.bar(x - width / 2, average_salary_year, width, label="средняя з/п")
        firstplt.bar(x + width / 2, average_salary_year_selected, width, label=f"з/п {self.vacancy_name}")
        firstplt.legend(fontsize=8)

    def generate_second_plot(self, secondplt):
        """Генерирует второй график

        Args: secondplt (axis) : второй график
        """
        labels = list(self.vacancies_count_by_year.keys())
        count_by_year = list(self.vacancies_count_by_year.values())
        count_by_year_selected = list(self.vacancies_count_by_year_selected_name.values())
        x = np.array(labels)
        y = np.array(count_by_year)
        width = 0.4
        secondplt.set_title("Количество вакансий по годам")
        secondplt.set_xticks(x)
        secondplt.tick_params(axis='x', rotation=90, labelsize=8)
        secondplt.set_yticks(np.arange(0, max(count_by_year), step=20000))
        secondplt.tick_params(axis='y', labelsize=8)
        secondplt.grid(axis='y')
        secondplt.bar(x - width / 2, count_by_year, width, label="Количество вакансий")
        secondplt.bar(x + width / 2, count_by_year_selected, width, label=f"Количество вакансий\n{self.vacancy_name}")
        secondplt.legend(fontsize=8)

    def generate_third_plot(self, thirdplt):
        """Генерирует второй график

        Args: thirdplt (axis) : третий график
        """
        y_cities = list(self.vacancies_top_average_salary_by_city.keys())
        for i in range(len(y_cities)):
            if "-" in y_cities[i]:
                y_cities[i] = "-\n".join(y_cities[i].split("-"))
        x_salaries = list(self.vacancies_top_average_salary_by_city.values())
        y = np.arange(len(y_cities))
        thirdplt.set_title("Уровень зарплат по городам")
        thirdplt.barh(y, x_salaries)
        thirdplt.set_yticks(y)
        thirdplt.set_yticklabels(y_cities, ha="right", va="center")
        thirdplt.tick_params(axis='y', labelsize=6)
        thirdplt.tick_params(axis='x', labelsize=8)
        thirdplt.invert_yaxis()
        thirdplt.grid(axis='x')

    def generate_fourth_plot(self, fourthplt):
        """Генерирует второй график

        Args: fourthplt (axis) : четвертый график
        """
        fourthplt.set_title("Доля вакансий по городам")
        others = 1
        for val in self.vacancies_fraction_by_city.values():
            others -= val
        fractions_by_city = dict(self.vacancies_fraction_by_city)
        fractions_by_city["Другие"] = others
        fractions_by_city = {k: v for k, v in sorted(fractions_by_city.items(),
                                                     reverse=True,
                                                     key=lambda item: item[1])}
        labels = list(fractions_by_city.keys())
        fractions = list(fractions_by_city.values())
        fourthplt.pie(fractions, labels=labels, textprops={'fontsize': 6})

    def generate_pdf(self):
        """Генерирует pdf файл"""
        env = Environment(loader=FileSystemLoader('.'))
        template = env.get_template("pdf_template.html")
        vacancy_name = self.vacancy_name
        book = load_workbook("report.xlsx")
        sheet = book.active
        sheet2 = book[(book.sheetnames)[1]]
        pdf_template = template.render({'vacancy_name': vacancy_name}, sheet=sheet, sheet2=sheet2)
        config = pdfkit.configuration(wkhtmltopdf=r'D:\wkhtmltox\bin\wkhtmltopdf.exe')
        pdfkit.from_string(pdf_template, 'report.pdf', configuration=config, options={"enable-local-file-access": None})


# report = Report(processed_data)
# report.generate_excel()
# report.generate_image()
# report.generate_pdf()

# def re_parse_year(date_string):
#     return re.findall(r'(\d{4})-(\d{1,2})-(\d{1,2})', date_string)[0][0]
#
#
# def parse_datetime(date_string):
#     """Парсер даты.
#
#     Returns:
#         datetime: Дата полученная с помощью strptime
#     """
#     return datetime.strptime(date_string, "%Y-%m-%dT%H:%M:%S%z")
#
#
# testData = DataSet("vacancies_by_year.csv").vacancies_objects
#
#
# def slice_parse_year(date_string):
#     return date_string[0:4]
#
#
# def parse_datetimes_from_vacancies(dataset_vacancies, func):
#     for vacancy in dataset_vacancies:
#         func(vacancy.published_at)
#
#
# def test_datetime_parser(dataset_vacancies, func):
#     cProfile.runctx('parse_datetimes_from_vacancies(dataset_vacancies, func)',
#                     {'func': func, 'dataset_vacancies': dataset_vacancies,
#                      "parse_datetimes_from_vacancies": parse_datetimes_from_vacancies}, {}, None, "cumtime")
#
#
# print("strptime")
# test_datetime_parser(testData, parse_datetime)
# print("re")
# test_datetime_parser(testData, re_parse_year)
# print("slice")
# test_datetime_parser(testData, slice_parse_year)
# dataset_test = DataSet("./chunks/2007.csv")
# print(dataset_test.process_vacancies("Аналитик"))

# Возможно придется возвращать vacancies_count_by_city и vacancies_average_salary_by_city для каждого из процессов,
# а затем в итоговом сконкатенированном массиве проводить окончательную обработку по процентам.

def concat_vacancy_dictionaries(dictionaries):
    """Склеивает словари в один"""
    result = {k: v for d in dictionaries for k, v in d.items()}
    return {key:value for key, value in sorted(result.items(), key=lambda item: int(item[0]))}

if __name__ == "__main__":
    with Pool() as pool:
        dataSets = pool.imap_unordered(DataSet, glob.glob("./subprograms/chunks/*.csv"))
        dataSets_with_vacancy_name = [(x,"Аналитик") for x in dataSets]
        processed_vacancies = list(pool.starmap(DataSet.process_vacancies, dataSets_with_vacancy_name))
        vacancies_average_salary_by_year = concat_vacancy_dictionaries([item[0] for item in processed_vacancies])

        print(vacancies_average_salary_by_year)
        vacancies_count_by_year = concat_vacancy_dictionaries([item[1] for item in processed_vacancies])
        vacancies_average_salary_by_year_selected_name = concat_vacancy_dictionaries([item[2] for item in processed_vacancies])
        vacancies_count_by_year_selected_name = concat_vacancy_dictionaries([item[3] for item in processed_vacancies])
        print(vacancies_count_by_year)
        print(vacancies_average_salary_by_year_selected_name)
        print(vacancies_count_by_year_selected_name)
        vacancies_count_by_city = reduce(lambda x, y: x + y, [Counter(item[7]) for item in processed_vacancies])
        # Найти общую сумму зарплат по городам (многомилионную), найти int mean
        # vacancies_average_salary_by_city = reduce(lambda x, y: x + y, [Counter(item[8]) for item in processed_vacancies])
        
        
        city_count = sum(vacancies_count_by_city.values())
        one_percent = floor(city_count / 100)
        # print(vacancies_count_by_city)
        # print(vacancies_average_salary_by_city)
        vacancies_fraction_by_city = {k: round(v / city_count, 4)
                                      for k, v in sorted(vacancies_count_by_city.items(),
                                                         reverse=True,
                                                         key=lambda item: item[1])
                                      if v >= one_percent}
        
        print(vacancies_fraction_by_city)
        vacancies_top_average_salary_by_city = {k: v
                                                for k, v in sorted(vacancies_average_salary_by_city.items(),
                                                                   reverse=True,
                                                                   key=lambda item: item[1])
                                                if vacancies_count_by_city[k] >= one_percent}
        print(vacancies_top_average_salary_by_city)
        # vacancies_top_average_salary_by_city = []
        # vacancies_fraction_by_city = []
        # print(vacancies_average_salary_by_year)
        # first = functools.reduce(lambda x, y: y[0].update(x[0]), processed_vacancies)
        # first = functools.reduce(lambda x, y: Counter(x[0]) + Counter(y[0]), processed_vacancies)

        
    





