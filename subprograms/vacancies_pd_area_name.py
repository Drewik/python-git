from openpyxl import Workbook
from openpyxl.styles import Border, Side, Font
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
from math import floor
from jinja2 import Environment, FileSystemLoader
import pdfkit
from openpyxl.styles.numbers import FORMAT_PERCENTAGE_00


# file_name = "vacancies_dif_currencies_with_salary.csv"
# # vacancy_name = "Аналитик"
# # area_name = "Москва"
file_name = input("Введите название файла: ")
vacancy_name = input("Введите название профессии: ")
area_name = input("Введите название региона: ")

df = pd.read_csv(file_name, encoding="utf-8")
q = df['salary'].quantile(0.99)
one_percent = df.shape[0] // 100
df_value_counts = df['area_name'].value_counts()
df_fraction = df['area_name'].value_counts(normalize=True)
vacancies_fraction_by_city = df_fraction.head(10).to_dict()
df_filtered_by_value_counts_threshold = df[df['area_name'].isin(df_value_counts[df_value_counts > one_percent].index)]
df_highest_average_salary_by_city = df_filtered_by_value_counts_threshold\
    .groupby('area_name')['salary']\
        .mean()\
            .astype(int)\
                .sort_values(ascending=False).head(10)
vacancies_top_average_salary_by_city = df_highest_average_salary_by_city.to_dict()

df = df[df['salary'] < q]
df = df[df['area_name'] == area_name]

def split_years(dt):
    return [dt[dt['published_at'].str[:4] == y] for y in dt['published_at'].str[:4].unique()]

res = split_years(df)

res_selected_vacancy = [i[i["name"].str.contains(vacancy_name)] for i in res]

vacancies_average_salary_by_year_selected_name = {}
vacancies_count_by_year_selected_name = {}
for i in res_selected_vacancy:
    vacancies_average_salary_by_year_selected_name[int(i['published_at'].iat[0][:4])] = floor(i['salary'].mean())
    vacancies_count_by_year_selected_name[int(i['published_at'].iat[0][:4])] = len(i.index)

print(vacancies_average_salary_by_year_selected_name)
print(vacancies_count_by_year_selected_name)


border = Border(left=Side(border_style='thin', color='000000'),
                right=Side(border_style='thin', color='000000'),
                top=Side(border_style='thin', color='000000'),
                bottom=Side(border_style='thin', color='000000'))
font = Font(bold=True)

def generate_excel():
    wb = Workbook()
    ws = wb.active
    ws.title = "Статистика по годам"
    ws2 = wb.create_sheet("Статистика по городам")
    fill_first_table(ws)
    set_correct_cell_width(ws)
    fill_second_table(ws2)
    set_correct_cell_width(ws2)
    return wb


def fill_first_table(ws):
    """Заполняет первый лист таблицы.
    Args:
        ws (openpyxl.workbook.workbook.Workbook @property): лист таблицы
    """
    set_first_page_headers(ws)
    years = list(vacancies_average_salary_by_year_selected_name.keys())
    for row in range(0, len(years)):
        set_cell_border(ws.cell(column=1, row=row + 2, value=years[row]))
        set_cell_border(
            ws.cell(column=2, row=row + 2, value=vacancies_average_salary_by_year_selected_name[years[row]]))
        set_cell_border(
            ws.cell(column=3, row=row + 2, value=vacancies_count_by_year_selected_name[years[row]]))

def fill_second_table(ws2):
    """Заполняет второй лист таблицы.
    Args:
        ws2 (openpyxl.workbook.workbook.Workbook @property): лист таблицы
    """
    set_second_page_headers(ws2)
    city_list_top_salary = list(vacancies_top_average_salary_by_city)
    for row in range(0, len(vacancies_top_average_salary_by_city)):
        set_cell_border(ws2.cell(column=1, row=row + 2, value=city_list_top_salary[row]))
        set_cell_border(ws2.cell(column=2, row=row + 2,
                                        value=vacancies_top_average_salary_by_city[city_list_top_salary[row]]))
    city_list_fraction = list(vacancies_fraction_by_city)
    for row in range(0, len(vacancies_top_average_salary_by_city)):
        set_cell_border(ws2.cell(column=4, row=row + 2, value=city_list_fraction[row]))
        cell = ws2.cell(column=5, row=row + 2, value=vacancies_fraction_by_city[city_list_fraction[row]])
        cell.number_format = FORMAT_PERCENTAGE_00
        set_cell_border(cell)

def set_correct_cell_width(ws):
    """Форматирует ширину клеток таблицы.
    Args:
        ws (openpyxl.workbook.workbook.Workbook @property): лист таблицы
    """
    dims = {}
    for row in ws.rows:
        for cell in row:
            if cell.value:
                dims[cell.column_letter] = max((dims.get(cell.column_letter, 0), len(str(cell.value))))
    for col, value in dims.items():
        ws.column_dimensions[col].width = value + 2

def set_cell_border(cell):
    """Устанавливает границу клетки таблицы.

    Args:
        cell ((openpyxl.workbook.workbook.Workbook @property).cell): Клетка таблицы
    """
    cell.border = border

def set_first_page_headers( ws):
    """Устанавливает заголовки первого листа таблицы.

    Args:
        ws (openpyxl.workbook.workbook.Workbook @property): Лист таблицы
    """
    ws["A1"] = "Год"
    ws["B1"] = "Средняя зарплата"
    ws["C1"] = "Количество вакансий"
    for row in ws.iter_rows(max_row=1, max_col=3):
        for cell in row:
            set_cell_border(cell)
            cell.font = font
def set_second_page_headers(ws2):
    """Устанавливает заголовки второго листа таблицы

    Args:
        ws2 (openpyxl.workbook.workbook.Workbook @property): Лист таблицы
    """
    ws2["A1"] = "Город"
    ws2["A1"].font = font
    ws2["A1"].border = border

    ws2["B1"] = "Уровень зарплат"
    ws2["B1"].font = font
    ws2["B1"].border = border

    ws2["D1"] = "Город"
    ws2["D1"].font = font
    ws2["D1"].border = border

    ws2["E1"] = "Доля вакансий"
    ws2["E1"].font = font
    ws2["E1"].border = border

def generate_image():
    """Генерирует изображение графиков"""
    fig, axes = plt.subplots(nrows=2, ncols=2)
    generate_first_plot(axes[0, 0])
    generate_second_plot(axes[0, 1])
    generate_third_plot(axes[1, 0])
    generate_fourth_plot(axes[1, 1])
    plt.tight_layout()
    plt.savefig(r"D:\PythonProject\report\pd_graph_area_name.png", dpi=200)

def generate_first_plot(firstplt):
    """Генерирует первый график

    Args: firstplt (axis): Первый график
    """
    labels = list(vacancies_average_salary_by_year_selected_name.keys())
    average_salary_year_selected = list(vacancies_average_salary_by_year_selected_name.values())
    x = np.array(labels)
    width = 0.4
    firstplt.set_title("Уровень зарплат по годам")
    firstplt.set_xticks(x)
    firstplt.tick_params(axis='x', rotation=90, labelsize=8)
    firstplt.set_yticks(np.arange(0, max(average_salary_year_selected), step=20000))
    firstplt.tick_params(axis='y', labelsize=8)
    firstplt.grid(axis="y")
    firstplt.bar(x, average_salary_year_selected, width, label=f"з/п {vacancy_name}")
    firstplt.legend(fontsize=8)

def generate_second_plot(secondplt):
    """Генерирует второй график

    Args: secondplt (axis) : второй график
    """
    labels = list(vacancies_count_by_year_selected_name.keys())
    count_by_year_selected = list(vacancies_count_by_year_selected_name.values())
    x = np.array(labels)
    y = np.array(count_by_year_selected)
    width = 0.4
    secondplt.set_title("Количество вакансий по годам")
    secondplt.set_xticks(x)
    secondplt.tick_params(axis='x', rotation=90, labelsize=8)
    secondplt.set_yticks(np.arange(0, max(count_by_year_selected),step=100))
    secondplt.tick_params(axis='y', labelsize=8)
    secondplt.grid(axis='y')
    secondplt.bar(x, count_by_year_selected, width, label=f"Количество вакансий\n{vacancy_name}")
    secondplt.legend(fontsize=8)

def generate_third_plot(thirdplt):
    """Генерирует второй график

    Args: thirdplt (axis) : третий график
    """
    y_cities = list(vacancies_top_average_salary_by_city.keys())
    for i in range(len(y_cities)):
        if "-" in y_cities[i]:
            y_cities[i] = "-\n".join(y_cities[i].split("-"))
    x_salaries = list(vacancies_top_average_salary_by_city.values())
    y = np.arange(len(y_cities))
    thirdplt.set_title("Уровень зарплат по городам")
    thirdplt.barh(y, x_salaries)
    thirdplt.set_yticks(y)
    thirdplt.set_yticklabels(y_cities, ha="right", va="center")
    thirdplt.tick_params(axis='y', labelsize=6)
    thirdplt.tick_params(axis='x', labelsize=8)
    thirdplt.invert_yaxis()
    thirdplt.grid(axis='x')

def generate_fourth_plot(fourthplt):
    """Генерирует второй график

    Args: fourthplt (axis) : четвертый график
    """
    fourthplt.set_title("Доля вакансий по городам")
    others = 1
    for val in vacancies_fraction_by_city.values():
        others -= val
    fractions_by_city = dict(vacancies_fraction_by_city)
    fractions_by_city["Другие"] = others
    fractions_by_city = {k: v for k, v in sorted(fractions_by_city.items(),
                                                    reverse=True,
                                                    key=lambda item: item[1])}
    labels = list(fractions_by_city.keys())
    fractions = list(fractions_by_city.values())
    fourthplt.pie(fractions, labels=labels, textprops={'fontsize': 6})

generate_image()

def generate_pdf():
    """Генерирует pdf файл"""
    env = Environment(loader=FileSystemLoader('.'))
    template = env.get_template("pdf_template_pd_area_name.html")
    book = generate_excel()
    sheet = book.active
    sheet2 = book[(book.sheetnames)[1]]
    pdf_template = template.render({'vacancy_name': vacancy_name}, sheet=sheet, sheet2=sheet2, area_name=area_name)
    config = pdfkit.configuration(wkhtmltopdf=r'D:\wkhtmltox\bin\wkhtmltopdf.exe')
    pdfkit.from_string(pdf_template, 'report_pd_area_name.pdf', configuration=config, options={"enable-local-file-access": None})

generate_pdf()