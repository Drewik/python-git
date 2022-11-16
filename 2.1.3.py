import csv
import re
from openpyxl import Workbook
from openpyxl.reader.excel import load_workbook
from openpyxl.styles import Border, Side, Font
from datetime import datetime
from math import floor
from statistics import mean
import matplotlib.pyplot as plt
import numpy as np
from openpyxl.styles.numbers import FORMAT_PERCENTAGE_00
from jinja2 import Environment, FileSystemLoader
import pdfkit


class Salary:
    currency_to_rub = {
        "AZN": 35.68,
        "BYR": 23.91,
        "EUR": 59.90,
        "GEL": 21.74,
        "KGS": 0.76,
        "KZT": 0.13,
        "RUR": 1,
        "UAH": 1.64,
        "USD": 60.66,
        "UZS": 0.0055,
    }

    def __init__(self, salary_from, salary_to, salary_currency):
        self.salary_from = salary_from
        self.salary_to = salary_to
        self.salary_currency = salary_currency
        if (self.salary_currency == "RUR"):
            self.rub_salary = floor(((float(self.salary_from) + float(self.salary_to)) / 2))
        else:
            self.rub_salary = floor((float(self.salary_from) * self.currency_to_rub[self.salary_currency]
                                     + float(self.salary_to) * self.currency_to_rub[self.salary_currency])
                                    / 2)


class Vacancy:
    def __init__(self, vacancy_dictionary):
        self.name = vacancy_dictionary.get("name")
        self.description = vacancy_dictionary.get("description")
        self.key_skills = vacancy_dictionary.get("key_skills")
        self.experience_id = vacancy_dictionary.get("experience_id")
        self.premium = vacancy_dictionary.get("premium")
        self.employer_name = vacancy_dictionary.get("employer_name")
        self.salary = Salary(vacancy_dictionary.get("salary_from"),
                             vacancy_dictionary.get("salary_to"),
                             vacancy_dictionary.get("salary_currency"))
        self.area_name = vacancy_dictionary.get("area_name")
        self.published_at = vacancy_dictionary.get("published_at")

    def parse_datetime(self):
        return datetime.strptime(self.published_at, "%Y-%m-%dT%H:%M:%S%z")


class DataSet:
    def __init__(self, file_name):
        self.file_name = file_name
        self.vacancies_objects = self.csv_OOP_parser(file_name)

    def remove_html(self, string):
        regex = re.compile(r'<[^>]+>')
        return regex.sub('', string)

    def repair_string(self, s):
        s = self.remove_html(s)
        s = s.strip()
        s = " ".join(s.split())
        return s

    def fill_vacancy_dictionary(self, field_names, row):
        vacancy_dictionary = {}
        if len(row) < len(field_names):
            return 0
        for i in range(len(field_names)):
            if len(row[i]) == 0:
                return 0
            self.repair_string(row[i])
            vacancy_dictionary[field_names[i]] = row[i]
        return vacancy_dictionary

    def csv_OOP_parser(self, file_name):
        f = open(file_name, mode="r", encoding="utf-8-sig")
        reader = csv.reader(f)
        try:
            field_names = next(reader)
        except:
            return 0

        vacancies_objects = []
        for row in reader:
            if len(row) < len(field_names):
                continue
            vacancy_dict = self.fill_vacancy_dictionary(field_names, row)
            if vacancy_dict == 0:
                continue
            vacancy = Vacancy(vacancy_dict)
            vacancies_objects.append(vacancy)
        return vacancies_objects

    def process_vacancies(self, vacancy_name):
        # 1. Динамика уровня зарплат по годам
        vacancies_average_salary_by_year = {}
        # 2. Динамика количества вакансий по годам
        vacancies_count_by_year = {}
        # 3. Динамика уровня зарплат по годам для выбранной профессии
        vacancies_average_salary_by_year_selected_name = {}
        # 4. Динамика количества вакансий по годам для выбранной профессии
        vacancies_count_by_year_selected_name = {}
        # 5. Уровень зарплат по городам (в порядке убывания) - только первые 10 значений
        vacancies_average_salary_by_city = {}
        vacancies_count_by_city = {}
        vacancies_objects = self.vacancies_objects
        for vacancy in vacancies_objects:
            year = int(vacancy.published_at[0:4])

            if vacancy.area_name in vacancies_count_by_city:
                vacancies_count_by_city[vacancy.area_name] += 1
            else:
                vacancies_count_by_city[vacancy.area_name] = 1

            if year in vacancies_count_by_year:
                vacancies_count_by_year[year] += 1
            else:
                vacancies_count_by_year[year] = 1
                vacancies_count_by_year_selected_name[year] = 0

            if year in vacancies_average_salary_by_year:
                vacancies_average_salary_by_year[year].append(vacancy.salary.rub_salary)
            else:
                vacancies_average_salary_by_year[year] = [vacancy.salary.rub_salary]
                vacancies_average_salary_by_year_selected_name[year] = 0

            if vacancy.area_name in vacancies_average_salary_by_city:
                vacancies_average_salary_by_city[vacancy.area_name].append(vacancy.salary.rub_salary)
            else:
                vacancies_average_salary_by_city[vacancy.area_name] = [vacancy.salary.rub_salary]

            if vacancy_name in vacancy.name:
                if year in vacancies_count_by_year_selected_name:
                    vacancies_count_by_year_selected_name[year] += 1

                if vacancies_average_salary_by_year_selected_name[year] == 0:
                    vacancies_average_salary_by_year_selected_name[year] = [vacancy.salary.rub_salary]
                else:
                    vacancies_average_salary_by_year_selected_name[year].append(vacancy.salary.rub_salary)

        for key in vacancies_average_salary_by_year:
            vacancies_average_salary_by_year[key] = int(mean(vacancies_average_salary_by_year[key]))
        for key in vacancies_average_salary_by_year_selected_name:
            if vacancies_average_salary_by_year_selected_name[key] != 0:
                vacancies_average_salary_by_year_selected_name[key] = int(
                    mean(vacancies_average_salary_by_year_selected_name[key]))
        for key in vacancies_average_salary_by_city:
            vacancies_average_salary_by_city[key] = int(mean(vacancies_average_salary_by_city[key]))

        city_count = sum(vacancies_count_by_city.values())
        one_percent = floor(city_count / 100)

        vacancies_fraction_by_city = {k: round(v / city_count, 4)
                                      for k, v in sorted(vacancies_count_by_city.items(),
                                                         reverse=True,
                                                         key=lambda item: item[1])
                                      if v >= one_percent}
        vacancies_top_average_salary_by_city = {k: v
                                                for k, v in sorted(vacancies_average_salary_by_city.items(),
                                                                   reverse=True,
                                                                   key=lambda item: item[1])
                                                if vacancies_count_by_city[k] >= one_percent}

        return vacancies_average_salary_by_year, \
               vacancies_count_by_year, \
               vacancies_average_salary_by_year_selected_name, \
               vacancies_count_by_year_selected_name, \
               vacancies_top_average_salary_by_city, \
               vacancies_fraction_by_city, \
               vacancy_name


file_name = input("Введите название файла: ")
vacancy_name = input("Введите название профессии: ")

dataset = DataSet(file_name)
processed_data = dataset.process_vacancies(vacancy_name)

print(f"Динамика уровня зарплат по годам: {processed_data[0]}")
print(f"Динамика количества вакансий по годам: {processed_data[1]}")
print(f"Динамика уровня зарплат по годам для выбранной профессии: {processed_data[2]}")
print(f"Динамика количества вакансий по годам для выбранной профессии: {processed_data[3]}")
print(f"Уровень зарплат по городам (в порядке убывания): {dict(list(processed_data[4].items())[:10])}")
print(f"Доля вакансий по городам (в порядке убывания): {dict(list(processed_data[5].items())[:10])}")


class Report:
    border = Border(left=Side(border_style='thin', color='000000'),
                    right=Side(border_style='thin', color='000000'),
                    top=Side(border_style='thin', color='000000'),
                    bottom=Side(border_style='thin', color='000000'))
    font = Font(bold=True)

    def __init__(self, processed_data):
        self.vacancies_average_salary_by_year = processed_data[0]
        self.vacancies_count_by_year = processed_data[1]
        self.vacancies_average_salary_by_year_selected_name = processed_data[2]
        self.vacancies_count_by_year_selected_name = processed_data[3]
        self.vacancies_top_average_salary_by_city = dict(list(processed_data[4].items())[:10])
        self.vacancies_fraction_by_city = dict(list(processed_data[5].items())[:10])
        self.vacancy_name = processed_data[6]

    def generate_excel(self):
        wb = Workbook()
        ws = wb.active
        ws.title = "Статистика по годам"
        ws2 = wb.create_sheet("Статистика по городам")
        self.fill_first_table(ws)
        self.set_correct_cell_width(ws)
        self.fill_second_table(ws2)
        self.set_correct_cell_width(ws2)
        ws2.column_dimensions['C'].width = 2
        wb.save("report.xlsx")

    def fill_first_table(self, ws):
        self.set_first_page_headers(ws)
        years = list(self.vacancies_average_salary_by_year.keys())
        for row in range(0, len(years)):
            self.set_cell_border(ws.cell(column=1, row=row + 2, value=years[row]))
            self.set_cell_border(
                ws.cell(column=2, row=row + 2, value=self.vacancies_average_salary_by_year[years[row]]))
            self.set_cell_border(
                ws.cell(column=3, row=row + 2, value=self.vacancies_average_salary_by_year_selected_name[years[row]]))
            self.set_cell_border(ws.cell(column=4, row=row + 2, value=self.vacancies_count_by_year[years[row]]))
            self.set_cell_border(
                ws.cell(column=5, row=row + 2, value=self.vacancies_count_by_year_selected_name[years[row]]))

    def fill_second_table(self, ws2):
        self.set_second_page_headers(ws2)
        city_list_top_salary = list(self.vacancies_top_average_salary_by_city)
        for row in range(0, len(self.vacancies_top_average_salary_by_city)):
            self.set_cell_border(ws2.cell(column=1, row=row + 2, value=city_list_top_salary[row]))
            self.set_cell_border(ws2.cell(column=2, row=row + 2,
                                          value=self.vacancies_top_average_salary_by_city[city_list_top_salary[row]]))
        city_list_fraction = list(self.vacancies_fraction_by_city)
        for row in range(0, len(self.vacancies_top_average_salary_by_city)):
            self.set_cell_border(ws2.cell(column=4, row=row + 2, value=city_list_fraction[row]))
            cell = ws2.cell(column=5, row=row + 2, value=self.vacancies_fraction_by_city[city_list_fraction[row]])
            cell.number_format = FORMAT_PERCENTAGE_00
            self.set_cell_border(cell)

    def set_correct_cell_width(self, ws):
        dims = {}
        for row in ws.rows:
            for cell in row:
                if cell.value:
                    dims[cell.column_letter] = max((dims.get(cell.column_letter, 0), len(str(cell.value))))
        for col, value in dims.items():
            ws.column_dimensions[col].width = value + 2

    def set_cell_border(self, cell):
        cell.border = self.border

    def set_first_page_headers(self, ws):
        ws["A1"] = "Год"
        ws["B1"] = "Средняя зарплата"
        ws["C1"] = f"Средняя зарплата - {self.vacancy_name}"
        ws["D1"] = "Количество вакансий"
        ws["E1"] = f"Количество вакансий - {self.vacancy_name}"
        for row in ws.iter_rows(max_row=1, max_col=5):
            for cell in row:
                self.set_cell_border(cell)
                cell.font = self.font

    def set_second_page_headers(self, ws2):
        ws2["A1"] = "Город"
        ws2["A1"].font = self.font
        ws2["A1"].border = self.border

        ws2["B1"] = "Уровень зарплат"
        ws2["B1"].font = self.font
        ws2["B1"].border = self.border

        ws2["D1"] = "Город"
        ws2["D1"].font = self.font
        ws2["D1"].border = self.border

        ws2["E1"] = "Доля вакансий"
        ws2["E1"].font = self.font
        ws2["E1"].border = self.border

    def generate_image(self):
        fig, axes = plt.subplots(nrows=2, ncols=2)
        self.generate_first_plot(axes[0, 0])
        self.generate_second_plot(axes[0, 1])
        self.generate_third_plot(axes[1, 0])
        self.generate_fourth_plot(axes[1, 1])
        plt.tight_layout()
        plt.savefig(r"D:\PythonProject\report\graph.png", dpi=200)

    def generate_first_plot(self, firstplt):
        labels = list(self.vacancies_average_salary_by_year.keys())
        average_salary_year = list(self.vacancies_average_salary_by_year.values())
        average_salary_year_selected = list(self.vacancies_average_salary_by_year_selected_name.values())
        x = np.array(labels)
        width = 0.4
        firstplt.set_title("Уровень зарплат по годам")
        firstplt.set_xticks(x)
        firstplt.tick_params(axis='x', rotation=90, labelsize=8)
        firstplt.set_yticks(np.arange(0, max(average_salary_year_selected), step=10000))
        firstplt.tick_params(axis='y', labelsize=8)
        firstplt.grid(axis="y")
        firstplt.bar(x - width / 2, average_salary_year, width, label="средняя з/п")
        firstplt.bar(x + width / 2, average_salary_year_selected, width, label=f"з/п {self.vacancy_name}")
        firstplt.legend(fontsize=8)

    def generate_second_plot(self, secondplt):
        labels = list(self.vacancies_count_by_year.keys())
        count_by_year = list(self.vacancies_count_by_year.values())
        count_by_year_selected = list(self.vacancies_count_by_year_selected_name.values())
        x = np.array(labels)
        y = np.array(count_by_year)
        width = 0.4
        secondplt.set_title("Количество вакансий по годам")
        secondplt.set_xticks(x)
        secondplt.tick_params(axis='x', rotation=90, labelsize=8)
        secondplt.set_yticks(np.arange(0, max(count_by_year), step=20000))
        secondplt.tick_params(axis='y', labelsize=8)
        secondplt.grid(axis='y')
        secondplt.bar(x - width / 2, count_by_year, width, label="Количество вакансий")
        secondplt.bar(x + width / 2, count_by_year_selected, width, label=f"Количество вакансий\n{self.vacancy_name}")
        secondplt.legend(fontsize=8)

    def generate_third_plot(self, thirdplt):
        y_cities = list(self.vacancies_top_average_salary_by_city.keys())
        for i in range(len(y_cities)):
            if "-" in y_cities[i]:
                y_cities[i] = "-\n".join(y_cities[i].split("-"))
        x_salaries = list(self.vacancies_top_average_salary_by_city.values())
        y = np.arange(len(y_cities))
        thirdplt.set_title("Уровень зарплат по городам")
        thirdplt.barh(y, x_salaries)
        thirdplt.set_yticks(y)
        thirdplt.set_yticklabels(y_cities, ha="right", va="center")
        thirdplt.tick_params(axis='y', labelsize=6)
        thirdplt.tick_params(axis='x', labelsize=8)
        thirdplt.invert_yaxis()
        thirdplt.grid(axis='x')

    def generate_fourth_plot(self, fourthplt):
        fourthplt.set_title("Доля вакансий по городам")
        others = 1
        for val in self.vacancies_fraction_by_city.values():
            others -= val
        fractions_by_city = dict(self.vacancies_fraction_by_city)
        fractions_by_city["Другие"] = others
        fractions_by_city = {k: v for k, v in sorted(fractions_by_city.items(),
                                                     reverse=True,
                                                     key=lambda item: item[1])}
        labels = list(fractions_by_city.keys())
        fractions = list(fractions_by_city.values())
        fourthplt.pie(fractions, labels=labels, textprops={'fontsize': 6})

    def generate_pdf(self):
        env = Environment(loader=FileSystemLoader('.'))
        template = env.get_template("pdf_template.html")
        vacancy_name = self.vacancy_name
        book = load_workbook("report.xlsx")
        sheet = book.active
        sheet2 = book[(book.sheetnames)[1]]
        pdf_template = template.render({'vacancy_name': vacancy_name}, sheet=sheet, sheet2=sheet2)
        config = pdfkit.configuration(wkhtmltopdf=r'D:\wkhtmltox\bin\wkhtmltopdf.exe')
        pdfkit.from_string(pdf_template, 'report.pdf', configuration=config, options={"enable-local-file-access": None})


report = Report(processed_data)
report.generate_excel()
report.generate_image()
report.generate_pdf()
